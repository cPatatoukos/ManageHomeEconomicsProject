from __future__ import annotations

from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from database import FinanceDB

MONTH_NAMES = [
    "Ιανουάριος", "Φεβρουάριος", "Μάρτιος", "Απρίλιος", "Μάιος", "Ιούνιος",
    "Ιούλιος", "Αύγουστος", "Σεπτέμβριος", "Οκτώβριος", "Νοέμβριος", "Δεκέμβριος",
]

HEADER_FILL = PatternFill("solid", fgColor="CCFBF1")
HEADER_FONT = Font(bold=True, color="0A0A0A")
TITLE_FONT = Font(bold=True, size=14)


def export_month_to_xlsx(
    db: FinanceDB,
    year: int,
    month: int,
    output_path: Path | str,
    *,
    user_id: Optional[int] = None,
) -> Path:
    path = Path(output_path)
    txns = db.list_transactions(year, month, user_id=user_id)
    totals = db.monthly_totals(year, month, user_id=user_id)
    family = user_id is None

    wb = Workbook()
    ws = wb.active
    ws.title = f"{MONTH_NAMES[month - 1][:3]} {year}"

    scope = "Οικογένεια" if family else "Προσωπικά"
    ws["A1"] = f"Οικονομική Αναφορά — {MONTH_NAMES[month - 1]} {year} ({scope})"
    ws["A1"].font = TITLE_FONT

    headers = ["ID", "Τίτλος", "Κατηγορία", "Τύπος", "Ποσό (€)", "Μηνιαίο"]
    if family:
        headers.insert(1, "Χρήστης")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].font = TITLE_FONT

    ws["A3"] = "Έσοδα"
    ws["B3"] = totals["income"]
    ws["A4"] = "Έξοδα"
    ws["B4"] = totals["expense"]
    ws["A5"] = "Καθαρό"
    ws["B5"] = totals["net"]

    row = 7
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for t in txns:
        row += 1
        values = [
            t["id"],
            t["title"],
            t["category_name"],
            "Έσοδο" if t["category_kind"] == "income" else "Έξοδο",
            float(t["amount"]),
            "Ναι" if t.get("recurring_template_id") else "Όχι",
        ]
        if family:
            values.insert(1, t.get("username", ""))
        for col, val in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=val)

    for col_idx in range(1, len(headers) + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
